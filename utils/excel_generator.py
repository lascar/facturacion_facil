# -*- coding: utf-8 -*-
"""
Generador de archivos Excel para informes
"""

import os
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path


class ExcelGenerator:
    """Clase para generar archivos Excel de informes"""
    
    def __init__(self):
        """Inicializar el generador de Excel"""
        self.workbook = None
        self.worksheet = None
    
    def generate_stock_excel(self, informe_data: Dict[str, Any], output_path: str) -> bool:
        """
        Genera un archivo Excel para el informe de stock
        
        Args:
            informe_data: Datos del informe de stock
            output_path: Ruta donde guardar el archivo Excel
            
        Returns:
            True si se generó correctamente, False en caso contrario
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Crear un nuevo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Informe de Stock"
            
            # Título
            ws['A1'] = "INFORME DE STOCK"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            
            # Fecha de generación
            ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:H2')
            
            # Encabezados
            headers = ['Nombre', 'Referencia', 'Categoría', 'Precio', 'Stock Actual', 'Stock Mínimo', 'Valor Stock', 'Última Actualización']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Datos de productos
            productos = informe_data.get('productos', [])
            for row_idx, producto in enumerate(productos, start=5):
                ws.cell(row=row_idx, column=1).value = producto.get('nombre', '')
                ws.cell(row=row_idx, column=2).value = producto.get('referencia', '')
                ws.cell(row=row_idx, column=3).value = producto.get('categoria', '')
                ws.cell(row=row_idx, column=4).value = producto.get('precio', 0.0)
                ws.cell(row=row_idx, column=4).number_format = '#,##0.00 €'
                
                stock_actual = producto.get('stock_actual', 0)
                stock_minimo = producto.get('stock_minimo', 0)
                
                # Stock actual con colorización
                cell_stock = ws.cell(row=row_idx, column=5)
                cell_stock.value = stock_actual
                cell_stock.alignment = Alignment(horizontal='center')
                
                if stock_actual == 0:
                    cell_stock.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell_stock.font = Font(color="FFFFFF", bold=True)
                elif stock_actual <= stock_minimo:
                    cell_stock.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Stock mínimo
                ws.cell(row=row_idx, column=6).value = stock_minimo
                ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center')
                
                # Valor stock
                ws.cell(row=row_idx, column=7).value = producto.get('valor_stock', 0.0)
                ws.cell(row=row_idx, column=7).number_format = '#,##0.00 €'
                
                # Última actualización
                ws.cell(row=row_idx, column=8).value = str(producto.get('fecha_actualizacion', 'N/A'))
            
            # Resumen
            resumen = informe_data.get('resumen', {})
            last_row = len(productos) + 6
            
            ws.cell(row=last_row, column=1).value = "RESUMEN"
            ws.cell(row=last_row, column=1).font = Font(bold=True, size=12)
            ws.merge_cells(f'A{last_row}:B{last_row}')
            
            ws.cell(row=last_row + 1, column=1).value = "Total Productos:"
            ws.cell(row=last_row + 1, column=2).value = resumen.get('total_productos', 0)
            
            ws.cell(row=last_row + 2, column=1).value = "Valor Total Stock:"
            ws.cell(row=last_row + 2, column=2).value = resumen.get('valor_total', 0.0)
            ws.cell(row=last_row + 2, column=2).number_format = '#,##0.00 €'
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20

            # Crear directorio si no existe usando pathlib
            output_file = Path(output_path)
            output_dir = output_file.parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Guardar el archivo
            wb.save(output_path)
            return True

        except ImportError:
            print("Error: openpyxl no está instalado. Instalar con: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Error generando Excel: {e}")
            return False

    def generate_facturacion_excel(self, informe_data: Dict[str, Any], output_path: str) -> bool:
        """
        Genera un archivo Excel para el informe de facturación

        Args:
            informe_data: Datos del informe de facturación
            output_path: Ruta donde guardar el archivo Excel

        Returns:
            True si se generó correctamente, False en caso contrario
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Crear un nuevo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Informe de Facturación"

            # Título
            ws['A1'] = "INFORME DE FACTURACIÓN"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:G1')

            # Fecha de generación
            fecha_inicio = informe_data.get('fecha_inicio', '')
            fecha_fin = informe_data.get('fecha_fin', '')
            ws['A2'] = f"Período: {fecha_inicio} - {fecha_fin}"
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:G2')

            current_row = 4

            # Lista de clientes
            clientes = informe_data.get('clientes', [])
            if clientes:
                ws.cell(row=current_row, column=1).value = "CLIENTES"
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1

                clientes_text = ', '.join([c.get('nombre', '') for c in clientes])
                ws.cell(row=current_row, column=1).value = clientes_text
                ws.merge_cells(f'A{current_row}:G{current_row}')
                current_row += 2

            # Facturas
            facturas = informe_data.get('facturas', [])
            if facturas:
                ws.cell(row=current_row, column=1).value = "FACTURAS"
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1

                # Encabezados de facturas
                headers = ['Número', 'Fecha', 'Cliente', 'Subtotal', 'IVA', 'Total', 'Estado']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1

                # Datos de facturas
                for factura in facturas:
                    ws.cell(row=current_row, column=1).value = factura.get('numero', '')
                    ws.cell(row=current_row, column=2).value = factura.get('fecha', '')
                    ws.cell(row=current_row, column=3).value = factura.get('cliente', '')
                    ws.cell(row=current_row, column=4).value = factura.get('subtotal', 0.0)
                    ws.cell(row=current_row, column=4).number_format = '#,##0.00 €'
                    ws.cell(row=current_row, column=5).value = factura.get('iva', 0.0)
                    ws.cell(row=current_row, column=5).number_format = '#,##0.00 €'
                    ws.cell(row=current_row, column=6).value = factura.get('total', 0.0)
                    ws.cell(row=current_row, column=6).number_format = '#,##0.00 €'
                    ws.cell(row=current_row, column=7).value = factura.get('estado', '')
                    current_row += 1

                current_row += 1

            # Resumen
            resumen = informe_data.get('resumen', {})
            ws.cell(row=current_row, column=1).value = "RESUMEN GENERAL"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1

            ws.cell(row=current_row, column=1).value = "Número de Facturas:"
            ws.cell(row=current_row, column=2).value = resumen.get('num_facturas', 0)
            current_row += 1

            # Desglose por IVA
            desglose_iva = informe_data.get('desglose_iva', [])
            if desglose_iva:
                ws.cell(row=current_row, column=1).value = "Desglose por IVA:"
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1

                for desglose in desglose_iva:
                    tasa = desglose.get('iva_aplicado', 0)
                    base = desglose.get('base_imponible', 0.0)
                    iva = desglose.get('total_iva', 0.0)
                    ws.cell(row=current_row, column=1).value = f"  IVA {tasa}%"
                    ws.cell(row=current_row, column=2).value = f"Base: {base:.2f} € + IVA: {iva:.2f} €"
                    current_row += 1

            current_row += 1
            ws.cell(row=current_row, column=1).value = "Total sin IVA:"
            ws.cell(row=current_row, column=2).value = resumen.get('subtotal', 0.0)
            ws.cell(row=current_row, column=2).number_format = '#,##0.00 €'
            current_row += 1

            ws.cell(row=current_row, column=1).value = "Total IVA:"
            ws.cell(row=current_row, column=2).value = resumen.get('total_iva', 0.0)
            ws.cell(row=current_row, column=2).number_format = '#,##0.00 €'
            current_row += 1

            ws.cell(row=current_row, column=1).value = "TOTAL CON IVA:"
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000", size=13)
            ws.cell(row=current_row, column=2).value = resumen.get('total', 0.0)
            ws.cell(row=current_row, column=2).number_format = '#,##0.00 €'
            ws.cell(row=current_row, column=2).font = Font(bold=True, color="FF0000", size=13)

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15

            # Crear directorio si no existe usando pathlib
            output_file = Path(output_path)
            output_dir = output_file.parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Guardar el archivo
            wb.save(output_path)
            return True

        except ImportError:
            print("Error: openpyxl no está instalado. Instalar con: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Error generando Excel: {e}")
            return False

