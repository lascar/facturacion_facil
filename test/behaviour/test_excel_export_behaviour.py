# -*- coding: utf-8 -*-
"""
Tests de comportamiento para la exportación de informes a Excel
"""

import os
import pytest
from datetime import datetime, timedelta
from database.database import Database
from services.informes_service import InformesService
from utils.excel_generator import ExcelGenerator


class TestExcelExportBehaviour:
    """Tests BDD para la exportación de informes a Excel"""
    
    @pytest.fixture(autouse=True)
    def setup(self, tmp_path):
        """Configuration pour chaque test"""
        # Créer une base de données temporaire
        self.db_path = str(tmp_path / "test_excel_export.db")
        self.db = Database(self.db_path)
        # La base de données est initialisée automatiquement dans __init__
        self.informes_service = InformesService(self.db_path)
        self.excel_generator = ExcelGenerator()
        self.output_dir = str(tmp_path / "output")
        os.makedirs(self.output_dir, exist_ok=True)
        
        yield
        
        # Nettoyage
        if os.path.exists(self.db_path):
            os.remove(self.db_path)
    
    def test_excel_stock_file_is_created(self):
        """
        GIVEN un informe de stock généré
        WHEN on exporte l'informe en Excel
        THEN un fichier Excel est créé
        """
        # GIVEN - Créer un produit
        producto_id = self.db.add_product({
            'nombre': 'Producto Test',
            'referencia': 'REF001',
            'precio': 10.0,
            'categoria': 'Test'
        })
        
        # Mettre à jour le stock
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE stock SET cantidad_disponible = ?, stock_minimo = ?
            WHERE producto_id = ?
        """, (5, 10, producto_id))
        conn.commit()
        conn.close()
        
        # Générer l'informe
        informe_data = self.informes_service.get_informe_stock()
        
        # WHEN - Exporter en Excel
        output_path = os.path.join(self.output_dir, "informe_stock.xlsx")
        success = self.excel_generator.generate_stock_excel(informe_data, output_path)
        
        # THEN - Le fichier existe
        assert success is True
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0
    
    def test_excel_stock_contains_product_data(self):
        """
        GIVEN un informe de stock avec plusieurs produits
        WHEN on exporte l'informe en Excel
        THEN le fichier Excel contient les données des produits
        """
        # GIVEN - Créer plusieurs produits
        productos_data = [
            {'nombre': 'Producto A', 'referencia': 'REF-A', 'precio': 10.0, 'categoria': 'Cat1', 'stock': 5, 'minimo': 10},
            {'nombre': 'Producto B', 'referencia': 'REF-B', 'precio': 20.0, 'categoria': 'Cat2', 'stock': 15, 'minimo': 5},
            {'nombre': 'Producto C', 'referencia': 'REF-C', 'precio': 30.0, 'categoria': 'Cat1', 'stock': 0, 'minimo': 3},
        ]

        for data in productos_data:
            producto_id = self.db.add_product({
                'nombre': data['nombre'],
                'referencia': data['referencia'],
                'precio': data['precio'],
                'categoria': data['categoria']
            })

            # Mettre à jour le stock avec une connexion séparée
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE stock SET cantidad_disponible = ?, stock_minimo = ?
                WHERE producto_id = ?
            """, (data['stock'], data['minimo'], producto_id))
            conn.commit()
            conn.close()
        
        # Générer l'informe
        informe_data = self.informes_service.get_informe_stock()
        
        # WHEN - Exporter en Excel
        output_path = os.path.join(self.output_dir, "informe_stock_multi.xlsx")
        success = self.excel_generator.generate_stock_excel(informe_data, output_path)
        
        # THEN - Le fichier existe et contient les données
        assert success is True
        assert os.path.exists(output_path)
        
        # Vérifier le contenu avec openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Vérifier le titre
        assert ws['A1'].value == "INFORME DE STOCK"
        
        # Vérifier les en-têtes (ligne 4)
        assert ws.cell(row=4, column=1).value == "Nombre"
        assert ws.cell(row=4, column=5).value == "Stock Actual"
        assert ws.cell(row=4, column=6).value == "Stock Mínimo"
        
        # Vérifier qu'il y a 3 produits (lignes 5, 6, 7)
        assert ws.cell(row=5, column=1).value is not None
        assert ws.cell(row=6, column=1).value is not None
        assert ws.cell(row=7, column=1).value is not None

    def test_excel_facturacion_file_is_created(self):
        """
        GIVEN un informe de facturación généré
        WHEN on exporte l'informe en Excel
        THEN un fichier Excel est créé
        """
        # GIVEN - Créer un client
        cliente_id = self.db.add_client({
            'nombre': 'Cliente Test',
            'dni': '12345678A',
            'email': 'test@test.com'
        })

        # Créer un produit
        producto_id = self.db.add_product({
            'nombre': 'Producto Test',
            'referencia': 'REF001',
            'precio': 100.0
        })

        # Créer une facture
        factura_data = {
            'numero': 'FACT-001',
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'cliente': {
                'id': cliente_id,
                'nombre': 'Cliente Test',
                'nif': '12345678A',
                'direccion': ''
            },
            'lineas': [
                {
                    'producto_id': producto_id,
                    'cantidad': 2,
                    'precio_unitario': 100.0,
                    'iva': 21.0,
                    'descuento': 0.0
                }
            ],
            'subtotal': 200.0,
            'iva_total': 42.0,
            'total': 242.0
        }

        factura_id = self.db.add_invoice(factura_data)

        # Générer l'informe
        fecha_inicio = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        fecha_fin = datetime.now().strftime('%Y-%m-%d')
        informe_data = self.informes_service.get_informe_facturacion(fecha_inicio, fecha_fin)

        # WHEN - Exporter en Excel
        output_path = os.path.join(self.output_dir, "informe_facturacion.xlsx")
        success = self.excel_generator.generate_facturacion_excel(informe_data, output_path)

        # THEN - Le fichier existe
        assert success is True
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_excel_facturacion_contains_invoice_data(self):
        """
        GIVEN un informe de facturación avec plusieurs factures
        WHEN on exporte l'informe en Excel
        THEN le fichier Excel contient les données des factures et le résumé
        """
        # GIVEN - Créer un client
        cliente_id = self.db.add_client({
            'nombre': 'Cliente Test',
            'dni': '12345678A',
            'email': 'test@test.com'
        })

        # Créer un produit
        producto_id = self.db.add_product({
            'nombre': 'Producto Test',
            'referencia': 'REF001',
            'precio': 100.0
        })

        # Créer plusieurs factures
        for i in range(3):
            cantidad = i + 1
            subtotal = cantidad * 100.0
            iva_total = subtotal * 0.21
            total = subtotal + iva_total

            factura_data = {
                'numero': f'FACT-00{i+1}',
                'fecha': datetime.now().strftime('%Y-%m-%d'),
                'cliente': {
                    'id': cliente_id,
                    'nombre': 'Cliente Test',
                    'nif': '12345678A',
                    'direccion': ''
                },
                'lineas': [
                    {
                        'producto_id': producto_id,
                        'cantidad': cantidad,
                        'precio_unitario': 100.0,
                        'iva': 21.0,
                        'descuento': 0.0
                    }
                ],
                'subtotal': subtotal,
                'iva_total': iva_total,
                'total': total
            }
            self.db.add_invoice(factura_data)

        # Générer l'informe
        fecha_inicio = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        fecha_fin = datetime.now().strftime('%Y-%m-%d')
        informe_data = self.informes_service.get_informe_facturacion(fecha_inicio, fecha_fin)

        # WHEN - Exporter en Excel
        output_path = os.path.join(self.output_dir, "informe_facturacion_multi.xlsx")
        success = self.excel_generator.generate_facturacion_excel(informe_data, output_path)

        # THEN - Le fichier existe et contient les données
        assert success is True
        assert os.path.exists(output_path)

        # Vérifier le contenu avec openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Vérifier le titre
        assert ws['A1'].value == "INFORME DE FACTURACIÓN"

        # Vérifier que le résumé contient le nombre de factures
        found_resumen = False
        for row in ws.iter_rows(min_row=1, max_row=50):
            if row[0].value == "RESUMEN GENERAL":
                found_resumen = True
                break

        assert found_resumen is True

